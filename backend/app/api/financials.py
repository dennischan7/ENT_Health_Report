"""
Financial data API routes.
Provides endpoints for querying balance sheets, income statements, and cash flow statements.
"""

from typing import Optional, List
from datetime import datetime
from io import BytesIO
import threading
from threading import Thread

from fastapi import APIRouter, Depends, HTTPException, status, Query
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from sqlalchemy import func, desc
from pydantic import BaseModel, Field

from app.db.session import get_db
from app.models.user import User
from app.models.enterprise import Enterprise
from app.models.financial import BalanceSheet, IncomeStatement, CashFlowStatement
from app.schemas.financial import (
    BalanceSheetResponse,
    BalanceSheetListResponse,
    IncomeStatementResponse,
    IncomeStatementListResponse,
    CashFlowStatementResponse,
    CashFlowStatementListResponse,
    EnterpriseFinancialSummary,
    EnterpriseFinancialSummaryList,
    EnterpriseFinancialDetail,
)
from app.api.deps import get_current_user


router = APIRouter()


# ==================== Balance Sheet Endpoints ====================


@router.get(
    "/balance-sheets",
    response_model=BalanceSheetListResponse,
    summary="List balance sheets",
)
def list_balance_sheets(
    enterprise_id: Optional[int] = Query(None, description="企业ID筛选"),
    company_code: Optional[str] = Query(None, description="公司代码筛选"),
    report_year: Optional[int] = Query(None, description="报告年度筛选"),
    page: int = Query(1, ge=1, description="页码"),
    page_size: int = Query(20, ge=1, le=100, description="每页数量"),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
) -> BalanceSheetListResponse:
    """
    List balance sheets with pagination and filters.
    支持按企业ID、公司代码、报告年度筛选。
    """
    query = db.query(BalanceSheet)

    # Apply filters
    if enterprise_id:
        query = query.filter(BalanceSheet.enterprise_id == enterprise_id)
    elif company_code:
        enterprise = db.query(Enterprise).filter(Enterprise.company_code == company_code).first()
        if enterprise:
            query = query.filter(BalanceSheet.enterprise_id == enterprise.id)
        else:
            return BalanceSheetListResponse(
                items=[], total=0, page=page, page_size=page_size, pages=0
            )

    if report_year:
        query = query.filter(BalanceSheet.report_year == report_year)

    # Count total
    total = query.count()

    # Paginate and order
    offset = (page - 1) * page_size
    items = (
        query.order_by(BalanceSheet.enterprise_id, desc(BalanceSheet.report_date))
        .offset(offset)
        .limit(page_size)
        .all()
    )

    return BalanceSheetListResponse(
        items=[BalanceSheetResponse.model_validate(item) for item in items],
        total=total,
        page=page,
        page_size=page_size,
        pages=(total + page_size - 1) // page_size,
    )


@router.get(
    "/balance-sheets/{balance_sheet_id}",
    response_model=BalanceSheetResponse,
    summary="Get balance sheet by ID",
)
def get_balance_sheet(
    balance_sheet_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
) -> BalanceSheetResponse:
    """Get a specific balance sheet by ID."""
    item = db.query(BalanceSheet).filter(BalanceSheet.id == balance_sheet_id).first()
    if not item:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail=f"Balance sheet with id {balance_sheet_id} not found",
        )
    return BalanceSheetResponse.model_validate(item)


# ==================== Income Statement Endpoints ====================


@router.get(
    "/income-statements",
    response_model=IncomeStatementListResponse,
    summary="List income statements",
)
def list_income_statements(
    enterprise_id: Optional[int] = Query(None, description="企业ID筛选"),
    company_code: Optional[str] = Query(None, description="公司代码筛选"),
    report_year: Optional[int] = Query(None, description="报告年度筛选"),
    page: int = Query(1, ge=1, description="页码"),
    page_size: int = Query(20, ge=1, le=100, description="每页数量"),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
) -> IncomeStatementListResponse:
    """
    List income statements with pagination and filters.
    支持按企业ID、公司代码、报告年度筛选。
    """
    query = db.query(IncomeStatement)

    # Apply filters
    if enterprise_id:
        query = query.filter(IncomeStatement.enterprise_id == enterprise_id)
    elif company_code:
        enterprise = db.query(Enterprise).filter(Enterprise.company_code == company_code).first()
        if enterprise:
            query = query.filter(IncomeStatement.enterprise_id == enterprise.id)
        else:
            return IncomeStatementListResponse(
                items=[], total=0, page=page, page_size=page_size, pages=0
            )

    if report_year:
        query = query.filter(IncomeStatement.report_year == report_year)

    # Count total
    total = query.count()

    # Paginate and order
    offset = (page - 1) * page_size
    items = (
        query.order_by(IncomeStatement.enterprise_id, desc(IncomeStatement.report_date))
        .offset(offset)
        .limit(page_size)
        .all()
    )

    return IncomeStatementListResponse(
        items=[IncomeStatementResponse.model_validate(item) for item in items],
        total=total,
        page=page,
        page_size=page_size,
        pages=(total + page_size - 1) // page_size,
    )


@router.get(
    "/income-statements/{income_statement_id}",
    response_model=IncomeStatementResponse,
    summary="Get income statement by ID",
)
def get_income_statement(
    income_statement_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
) -> IncomeStatementResponse:
    """Get a specific income statement by ID."""
    item = db.query(IncomeStatement).filter(IncomeStatement.id == income_statement_id).first()
    if not item:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail=f"Income statement with id {income_statement_id} not found",
        )
    return IncomeStatementResponse.model_validate(item)


# ==================== Cash Flow Statement Endpoints ====================


@router.get(
    "/cash-flow-statements",
    response_model=CashFlowStatementListResponse,
    summary="List cash flow statements",
)
def list_cash_flow_statements(
    enterprise_id: Optional[int] = Query(None, description="企业ID筛选"),
    company_code: Optional[str] = Query(None, description="公司代码筛选"),
    report_year: Optional[int] = Query(None, description="报告年度筛选"),
    page: int = Query(1, ge=1, description="页码"),
    page_size: int = Query(20, ge=1, le=100, description="每页数量"),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
) -> CashFlowStatementListResponse:
    """
    List cash flow statements with pagination and filters.
    支持按企业ID、公司代码、报告年度筛选。
    """
    query = db.query(CashFlowStatement)

    # Apply filters
    if enterprise_id:
        query = query.filter(CashFlowStatement.enterprise_id == enterprise_id)
    elif company_code:
        enterprise = db.query(Enterprise).filter(Enterprise.company_code == company_code).first()
        if enterprise:
            query = query.filter(CashFlowStatement.enterprise_id == enterprise.id)
        else:
            return CashFlowStatementListResponse(
                items=[], total=0, page=page, page_size=page_size, pages=0
            )

    if report_year:
        query = query.filter(CashFlowStatement.report_year == report_year)

    # Count total
    total = query.count()

    # Paginate and order
    offset = (page - 1) * page_size
    items = (
        query.order_by(CashFlowStatement.enterprise_id, desc(CashFlowStatement.report_date))
        .offset(offset)
        .limit(page_size)
        .all()
    )

    return CashFlowStatementListResponse(
        items=[CashFlowStatementResponse.model_validate(item) for item in items],
        total=total,
        page=page,
        page_size=page_size,
        pages=(total + page_size - 1) // page_size,
    )


@router.get(
    "/cash-flow-statements/{cash_flow_id}",
    response_model=CashFlowStatementResponse,
    summary="Get cash flow statement by ID",
)
def get_cash_flow_statement(
    cash_flow_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
) -> CashFlowStatementResponse:
    """Get a specific cash flow statement by ID."""
    item = db.query(CashFlowStatement).filter(CashFlowStatement.id == cash_flow_id).first()
    if not item:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail=f"Cash flow statement with id {cash_flow_id} not found",
        )
    return CashFlowStatementResponse.model_validate(item)


# ==================== Enterprise Financial Summary Endpoints ====================


@router.get(
    "/enterprises/summary",
    response_model=EnterpriseFinancialSummaryList,
    summary="List enterprises with financial data summary",
)
def list_enterprises_financial_summary(
    search: Optional[str] = Query(None, description="搜索公司代码或简称"),
    has_data: Optional[bool] = Query(None, description="是否有财务数据"),
    page: int = Query(1, ge=1, description="页码"),
    page_size: int = Query(20, ge=1, le=100, description="每页数量"),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
) -> EnterpriseFinancialSummaryList:
    """
    List enterprises with their financial data availability summary.
    显示每个企业的财务数据导入情况。
    """
    # Subqueries for counts
    balance_count_subq = (
        db.query(BalanceSheet.enterprise_id, func.count(BalanceSheet.id).label("count"))
        .group_by(BalanceSheet.enterprise_id)
        .subquery()
    )

    income_count_subq = (
        db.query(IncomeStatement.enterprise_id, func.count(IncomeStatement.id).label("count"))
        .group_by(IncomeStatement.enterprise_id)
        .subquery()
    )

    cashflow_count_subq = (
        db.query(CashFlowStatement.enterprise_id, func.count(CashFlowStatement.id).label("count"))
        .group_by(CashFlowStatement.enterprise_id)
        .subquery()
    )

    # Latest report date subquery
    latest_date_subq = (
        db.query(
            BalanceSheet.enterprise_id, func.max(BalanceSheet.report_date).label("latest_date")
        )
        .group_by(BalanceSheet.enterprise_id)
        .subquery()
    )

    # Main query
    query = (
        db.query(
            Enterprise.id.label("enterprise_id"),
            Enterprise.company_code,
            Enterprise.company_name,
            func.coalesce(balance_count_subq.c.count, 0).label("balance_sheet_count"),
            func.coalesce(income_count_subq.c.count, 0).label("income_statement_count"),
            func.coalesce(cashflow_count_subq.c.count, 0).label("cashflow_statement_count"),
            latest_date_subq.c.latest_date.label("latest_report_date"),
        )
        .outerjoin(balance_count_subq, Enterprise.id == balance_count_subq.c.enterprise_id)
        .outerjoin(income_count_subq, Enterprise.id == income_count_subq.c.enterprise_id)
        .outerjoin(cashflow_count_subq, Enterprise.id == cashflow_count_subq.c.enterprise_id)
        .outerjoin(latest_date_subq, Enterprise.id == latest_date_subq.c.enterprise_id)
    )

    # Apply filters
    if search:
        query = query.filter(
            (Enterprise.company_code.ilike(f"%{search}%"))
            | (Enterprise.company_name.ilike(f"%{search}%"))
        )

    if has_data is not None:
        if has_data:
            query = query.filter(balance_count_subq.c.count > 0)
        else:
            query = query.filter(func.coalesce(balance_count_subq.c.count, 0) == 0)

    # Count total
    total = query.count()

    # Paginate
    offset = (page - 1) * page_size
    results = query.order_by(Enterprise.company_code).offset(offset).limit(page_size).all()

    items = [
        EnterpriseFinancialSummary(
            enterprise_id=r.enterprise_id,
            company_code=r.company_code,
            company_name=r.company_name,
            balance_sheet_count=r.balance_sheet_count or 0,
            income_statement_count=r.income_statement_count or 0,
            cashflow_statement_count=r.cashflow_statement_count or 0,
            latest_report_date=r.latest_report_date,
        )
        for r in results
    ]

    return EnterpriseFinancialSummaryList(
        items=items,
        total=total,
        page=page,
        page_size=page_size,
        pages=(total + page_size - 1) // page_size,
    )


@router.get(
    "/enterprises/{enterprise_id}",
    response_model=EnterpriseFinancialDetail,
    summary="Get complete financial data for an enterprise",
)
def get_enterprise_financial_detail(
    enterprise_id: int,
    years: int = Query(5, ge=1, le=10, description="获取最近N年数据"),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
) -> EnterpriseFinancialDetail:
    """
    Get complete financial data (all three statements) for a specific enterprise.
    获取指定企业的完整财务数据（三大报表）。
    """
    enterprise = db.query(Enterprise).filter(Enterprise.id == enterprise_id).first()
    if not enterprise:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail=f"Enterprise with id {enterprise_id} not found",
        )

    # Get balance sheets
    balance_sheets = (
        db.query(BalanceSheet)
        .filter(BalanceSheet.enterprise_id == enterprise_id)
        .order_by(desc(BalanceSheet.report_date))
        .limit(years)
        .all()
    )

    # Get income statements
    income_statements = (
        db.query(IncomeStatement)
        .filter(IncomeStatement.enterprise_id == enterprise_id)
        .order_by(desc(IncomeStatement.report_date))
        .limit(years)
        .all()
    )

    # Get cash flow statements
    cash_flow_statements = (
        db.query(CashFlowStatement)
        .filter(CashFlowStatement.enterprise_id == enterprise_id)
        .order_by(desc(CashFlowStatement.report_date))
        .limit(years)
        .all()
    )

    return EnterpriseFinancialDetail(
        enterprise_id=enterprise.id,
        company_code=enterprise.company_code,
        company_name=enterprise.company_name,
        balance_sheets=[BalanceSheetResponse.model_validate(bs) for bs in balance_sheets],
        income_statements=[
            IncomeStatementResponse.model_validate(is_) for is_ in income_statements
        ],
        cash_flow_statements=[
            CashFlowStatementResponse.model_validate(cf) for cf in cash_flow_statements
        ],
    )


@router.get(
    "/enterprises/code/{company_code}",
    response_model=EnterpriseFinancialDetail,
    summary="Get complete financial data by company code",
)
def get_enterprise_financial_detail_by_code(
    company_code: str,
    years: int = Query(5, ge=1, le=10, description="获取最近N年数据"),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
) -> EnterpriseFinancialDetail:
    """
    Get complete financial data by company code.
    通过公司代码获取完整财务数据。
    """
    enterprise = db.query(Enterprise).filter(Enterprise.company_code == company_code).first()
    if not enterprise:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail=f"Enterprise with code {company_code} not found",
        )

    return get_enterprise_financial_detail(enterprise.id, years, db, current_user)


# ==================== Global Statistics Endpoint ====================


class GlobalFinancialStats(BaseModel):
    """Global financial data statistics."""

    total_enterprises: int = Field(..., description="企业总数")
    enterprises_with_data: int = Field(..., description="有财务数据的企业数")
    data_coverage_rate: float = Field(..., description="数据覆盖率(%)")
    balance_sheet_records: int = Field(..., description="资产负债表记录数")
    income_statement_records: int = Field(..., description="利润表记录数")
    cashflow_statement_records: int = Field(..., description="现金流量表记录数")
    total_records: int = Field(..., description="财务报表总记录数")


@router.get(
    "/stats",
    response_model=GlobalFinancialStats,
    summary="Get global financial data statistics",
)
def get_global_stats(
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
) -> GlobalFinancialStats:
    """
    Get global statistics for all financial data.
    获取财务数据全局统计信息。
    """
    # Total enterprises
    total_enterprises = db.query(func.count(Enterprise.id)).scalar()

    # Enterprises with data (have at least one balance sheet)
    enterprises_with_data = db.query(func.count(func.distinct(BalanceSheet.enterprise_id))).scalar()

    # Record counts
    balance_sheet_records = db.query(func.count(BalanceSheet.id)).scalar()
    income_statement_records = db.query(func.count(IncomeStatement.id)).scalar()
    cashflow_statement_records = db.query(func.count(CashFlowStatement.id)).scalar()

    # Calculate coverage rate
    data_coverage_rate = (
        round(enterprises_with_data / total_enterprises * 100, 1) if total_enterprises > 0 else 0
    )

    return GlobalFinancialStats(
        total_enterprises=total_enterprises,
        enterprises_with_data=enterprises_with_data,
        data_coverage_rate=data_coverage_rate,
        balance_sheet_records=balance_sheet_records,
        income_statement_records=income_statement_records,
        cashflow_statement_records=cashflow_statement_records,
        total_records=balance_sheet_records + income_statement_records + cashflow_statement_records,
    )


# ==================== Enterprise Data Status & Update ====================


class EnterpriseDataStatus(BaseModel):
    """Enterprise financial data status."""

    enterprise_id: int
    company_code: str
    company_name: str
    has_data: bool = Field(..., description="是否有财务数据")
    latest_year: Optional[int] = Field(None, description="最新数据年度")
    earliest_year: Optional[int] = Field(None, description="最早数据年度")
    total_years: int = Field(0, description="数据年数")
    expected_years: int = Field(5, description="期望年数")
    missing_years: list[int] = Field(default_factory=list, description="缺失的年度")
    need_update: bool = Field(False, description="是否需要更新")
    status: str = Field("no_data", description="状态: no_data/partial/complete")


@router.get(
    "/enterprises/{enterprise_id}/status",
    response_model=EnterpriseDataStatus,
    summary="Get enterprise financial data status",
)
def get_enterprise_data_status(
    enterprise_id: int,
    years: int = Query(5, ge=1, le=10, description="期望的数据年数"),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
) -> EnterpriseDataStatus:
    """
    Get the financial data status for a specific enterprise.
    获取指定企业的财务数据状态。
    """
    from datetime import datetime

    # Get enterprise
    enterprise = db.query(Enterprise).filter(Enterprise.id == enterprise_id).first()
    if not enterprise:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail=f"Enterprise with id {enterprise_id} not found",
        )

    # Get existing years from balance sheets
    existing_years = [
        row[0]
        for row in db.query(BalanceSheet.report_year)
        .filter(BalanceSheet.enterprise_id == enterprise_id)
        .distinct()
        .all()
    ]

    # Calculate expected years (current year and previous years)
    current_year = datetime.now().year
    expected_years_set = set(range(current_year - years + 1, current_year + 1))

    # Calculate missing years
    existing_set = set(existing_years)
    missing_years = sorted(expected_years_set - existing_set, reverse=True)

    # Determine status
    if not existing_years:
        data_status = "no_data"
        has_data = False
        need_update = True
    elif len(existing_years) < years:
        data_status = "partial"
        has_data = True
        need_update = len(missing_years) > 0
    else:
        data_status = "complete"
        has_data = True
        need_update = False

    return EnterpriseDataStatus(
        enterprise_id=enterprise_id,
        company_code=enterprise.company_code,
        company_name=enterprise.company_name,
        has_data=has_data,
        latest_year=max(existing_years) if existing_years else None,
        earliest_year=min(existing_years) if existing_years else None,
        total_years=len(existing_years),
        expected_years=years,
        missing_years=missing_years,
        need_update=need_update,
        status=data_status,
    )


@router.post(
    "/enterprises/{enterprise_id}/refresh",
    summary="Refresh enterprise financial data",
)
def refresh_enterprise_data(
    enterprise_id: int,
    years: int = Query(5, ge=1, le=10, description="获取最近N年数据"),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """
    Refresh financial data for a specific enterprise.
    Only fetches missing years of data.
    刷新指定企业的财务数据，仅获取缺失年度的数据。
    """
    from app.services.batch_import import BatchImportService

    # Get enterprise
    enterprise = db.query(Enterprise).filter(Enterprise.id == enterprise_id).first()
    if not enterprise:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail=f"Enterprise with id {enterprise_id} not found",
        )

    # Import data
    service = BatchImportService(db)
    result = service.import_enterprise_data_ths(enterprise, years=years, skip_existing=True)

    return {
        "enterprise_id": enterprise_id,
        "company_code": enterprise.company_code,
        "company_name": enterprise.company_name,
        "balance_count": result.get("balance_count", 0),
        "income_count": result.get("income_count", 0),
        "cashflow_count": result.get("cashflow_count", 0),
        "success": True,
        "message": f"成功更新 {result.get('balance_count', 0)} 条资产负债表、{result.get('income_count', 0)} 条利润表、{result.get('cashflow_count', 0)} 条现金流量表数据",
    }


# ==================== Batch Refresh Endpoints ====================

# Global state for batch refresh (in-memory, per-process)
_batch_refresh_state = {
    "is_running": False,
    "total": 0,
    "completed": 0,
    "failed": 0,
    "started_at": None,
    "current_enterprise": None,
    "lock": threading.Lock(),
}


class BatchRefreshStatus(BaseModel):
    """Batch refresh status response."""

    is_running: bool
    total: int
    completed: int
    failed: int
    started_at: Optional[datetime]
    current_enterprise: Optional[str]


def _run_batch_refresh(db_url: str, years: int):
    """Background task to refresh all enterprises."""
    from sqlalchemy import create_engine
    from sqlalchemy.orm import sessionmaker, scoped_session

    engine = create_engine(db_url, pool_pre_ping=True)
    SessionFactory = sessionmaker(bind=engine)
    Session = scoped_session(SessionFactory)
    db = Session()

    try:
        from app.models.enterprise import Enterprise
        from app.services.batch_import import BatchImportService

        enterprises = db.query(Enterprise).all()
        total = len(enterprises)

        _batch_refresh_state["total"] = total
        service = BatchImportService(db)

        for i, enterprise in enumerate(enterprises):
            if not _batch_refresh_state["is_running"]:
                break

            _batch_refresh_state["current_enterprise"] = (
                f"{enterprise.company_code} - {enterprise.company_name}"
            )

            try:
                result = service.import_enterprise_data_ths(
                    enterprise, years=years, skip_existing=True
                )
                if result.get("balance_count", 0) > 0 or result.get("income_count", 0) > 0:
                    _batch_refresh_state["completed"] += 1
                else:
                    _batch_refresh_state["failed"] += 1
            except Exception:
                _batch_refresh_state["failed"] += 1

            _batch_refresh_state["completed"] = i + 1

    finally:
        db.close()
        Session.remove()
        _batch_refresh_state["is_running"] = False
        _batch_refresh_state["current_enterprise"] = None


@router.get(
    "/batch-refresh/status",
    response_model=BatchRefreshStatus,
    summary="Get batch refresh status",
)
def get_batch_refresh_status(
    current_user: User = Depends(get_current_user),
) -> BatchRefreshStatus:
    """获取批量更新状态"""
    with _batch_refresh_state["lock"]:
        return BatchRefreshStatus(
            is_running=_batch_refresh_state["is_running"],
            total=_batch_refresh_state["total"],
            completed=_batch_refresh_state["completed"],
            failed=_batch_refresh_state["failed"],
            started_at=_batch_refresh_state["started_at"],
            current_enterprise=_batch_refresh_state["current_enterprise"],
        )


@router.post(
    "/batch-refresh/start",
    summary="Start batch refresh",
)
def start_batch_refresh(
    years: int = Query(5, ge=1, le=10, description="获取最近N年数据"),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """启动批量更新所有企业财务数据"""
    with _batch_refresh_state["lock"]:
        if _batch_refresh_state["is_running"]:
            return {
                "status": "already_running",
                "message": "批量更新正在进行中，请等待完成",
                "progress": {
                    "completed": _batch_refresh_state["completed"],
                    "total": _batch_refresh_state["total"],
                },
            }

        _batch_refresh_state["is_running"] = True
        _batch_refresh_state["total"] = db.query(Enterprise).count()
        _batch_refresh_state["completed"] = 0
        _batch_refresh_state["failed"] = 0
        _batch_refresh_state["started_at"] = datetime.now()
        _batch_refresh_state["current_enterprise"] = None

    from app.core.config import settings

    db_url = settings.DATABASE_URL

    thread = Thread(target=_run_batch_refresh, args=(db_url, years), daemon=True)
    thread.start()

    return {
        "status": "started",
        "message": f"批量更新已启动，共 {_batch_refresh_state['total']} 家企业",
        "total": _batch_refresh_state["total"],
    }


@router.post(
    "/batch-refresh/stop",
    summary="Stop batch refresh",
)
def stop_batch_refresh(
    current_user: User = Depends(get_current_user),
):
    """停止正在运行的批量更新"""
    with _batch_refresh_state["lock"]:
        if not _batch_refresh_state["is_running"]:
            return {
                "status": "not_running",
                "message": "当前没有正在运行的批量更新任务",
            }

        _batch_refresh_state["is_running"] = False

    return {
        "status": "stopped",
        "message": "批量更新已停止",
        "completed": _batch_refresh_state["completed"],
        "total": _batch_refresh_state["total"],
    }


# ==================== Batch Export Endpoints ====================


class ExportRequest(BaseModel):
    """Export request body."""

    enterprise_ids: List[int] = Field(..., description="企业ID列表")
    years: List[int] = Field(..., description="年份列表")


@router.post(
    "/export",
    summary="Export financial data to Excel",
)
def export_financial_data(
    request: ExportRequest,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """
    批量导出企业财务数据到Excel
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, Border, Side

    # Create workbook
    wb = Workbook()

    # Create sheets
    ws_balance = wb.active
    ws_balance.title = "资产负债表"
    ws_income = wb.create_sheet("利润表")
    ws_cashflow = wb.create_sheet("现金流量表")

    # Style
    header_font = Font(bold=True)
    header_alignment = Alignment(horizontal="center")
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    # Headers for balance sheet
    balance_headers = [
        "公司代码",
        "公司名称",
        "报告日期",
        "报告年度",
        "货币资金",
        "交易性金融资产",
        "应收账款",
        "存货",
        "流动资产合计",
        "固定资产",
        "资产总计",
        "短期借款",
        "应付账款",
        "流动负债合计",
        "长期借款",
        "负债合计",
        "实收资本",
        "未分配利润",
        "所有者权益合计",
    ]

    # Headers for income statement
    income_headers = [
        "公司代码",
        "公司名称",
        "报告日期",
        "报告年度",
        "营业收入",
        "营业成本",
        "销售费用",
        "管理费用",
        "财务费用",
        "营业利润",
        "利润总额",
        "所得税费用",
        "净利润",
        "归属母公司净利润",
        "基本每股收益",
    ]

    # Headers for cash flow statement
    cashflow_headers = [
        "公司代码",
        "公司名称",
        "报告日期",
        "报告年度",
        "销售商品收到的现金",
        "税费返还",
        "经营活动现金流入小计",
        "购买商品支付的现金",
        "支付给职工的现金",
        "支付的各项税费",
        "经营活动现金流出小计",
        "经营活动产生的现金流量净额",
        "收回投资收到的现金",
        "取得投资收益收到的现金",
        "投资活动现金流入小计",
        "购建固定资产支付的现金",
        "投资支付的现金",
        "投资活动现金流出小计",
        "投资活动产生的现金流量净额",
        "吸收投资收到的现金",
        "取得借款收到的现金",
        "筹资活动现金流入小计",
        "偿还债务支付的现金",
        "分配股利支付的现金",
        "筹资活动现金流出小计",
        "筹资活动产生的现金流量净额",
        "现金及现金等价物净增加额",
        "期末现金及现金等价物余额",
    ]

    # Write headers
    for col, header in enumerate(balance_headers, 1):
        cell = ws_balance.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.alignment = header_alignment
        cell.border = thin_border

    for col, header in enumerate(income_headers, 1):
        cell = ws_income.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.alignment = header_alignment
        cell.border = thin_border

    for col, header in enumerate(cashflow_headers, 1):
        cell = ws_cashflow.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.alignment = header_alignment
        cell.border = thin_border

    # Data rows
    row_balance = 2
    row_income = 2
    row_cashflow = 2

    for ent_id in request.enterprise_ids:
        enterprise = db.query(Enterprise).filter(Enterprise.id == ent_id).first()
        if not enterprise:
            continue

        # Get financial data for specified years
        for year in request.years:
            # Balance sheet
            balance = (
                db.query(BalanceSheet)
                .filter(BalanceSheet.enterprise_id == ent_id, BalanceSheet.report_year == year)
                .first()
            )

            if balance:
                row_data = [
                    enterprise.company_code,
                    enterprise.company_name,
                    str(balance.report_date),
                    balance.report_year,
                    float(balance.cash) if balance.cash else 0,
                    float(balance.trading_financial_assets)
                    if balance.trading_financial_assets
                    else 0,
                    float(balance.accounts_receivable) if balance.accounts_receivable else 0,
                    float(balance.inventory) if balance.inventory else 0,
                    float(balance.total_current_assets) if balance.total_current_assets else 0,
                    float(balance.fixed_assets) if balance.fixed_assets else 0,
                    float(balance.total_assets) if balance.total_assets else 0,
                    float(balance.short_term_borrowings) if balance.short_term_borrowings else 0,
                    float(balance.accounts_payable) if balance.accounts_payable else 0,
                    float(balance.total_current_liabilities)
                    if balance.total_current_liabilities
                    else 0,
                    float(balance.long_term_borrowings) if balance.long_term_borrowings else 0,
                    float(balance.total_liabilities) if balance.total_liabilities else 0,
                    float(balance.paid_in_capital) if balance.paid_in_capital else 0,
                    float(balance.retained_earnings) if balance.retained_earnings else 0,
                    float(balance.total_equity) if balance.total_equity else 0,
                ]
                for col, value in enumerate(row_data, 1):
                    ws_balance.cell(row=row_balance, column=col, value=value).border = thin_border
                row_balance += 1

            # Income statement
            income = (
                db.query(IncomeStatement)
                .filter(
                    IncomeStatement.enterprise_id == ent_id, IncomeStatement.report_year == year
                )
                .first()
            )

            if income:
                row_data = [
                    enterprise.company_code,
                    enterprise.company_name,
                    str(income.report_date),
                    income.report_year,
                    float(income.operating_revenue) if income.operating_revenue else 0,
                    float(income.operating_cost) if income.operating_cost else 0,
                    float(income.selling_expenses) if income.selling_expenses else 0,
                    float(income.admin_expenses) if income.admin_expenses else 0,
                    float(income.financial_expenses) if income.financial_expenses else 0,
                    float(income.operating_profit) if income.operating_profit else 0,
                    float(income.total_profit) if income.total_profit else 0,
                    float(income.income_tax) if income.income_tax else 0,
                    float(income.net_profit) if income.net_profit else 0,
                    float(income.net_profit_parent) if income.net_profit_parent else 0,
                    float(income.basic_eps) if income.basic_eps else 0,
                ]
                for col, value in enumerate(row_data, 1):
                    ws_income.cell(row=row_income, column=col, value=value).border = thin_border
                row_income += 1

            # Cash flow statement
            cashflow = (
                db.query(CashFlowStatement)
                .filter(
                    CashFlowStatement.enterprise_id == ent_id, CashFlowStatement.report_year == year
                )
                .first()
            )

            if cashflow:
                row_data = [
                    enterprise.company_code,
                    enterprise.company_name,
                    str(cashflow.report_date),
                    cashflow.report_year,
                    float(cashflow.cash_received_sales) if cashflow.cash_received_sales else 0,
                    float(cashflow.tax_refund_received) if cashflow.tax_refund_received else 0,
                    0,  # 经营活动现金流入小计
                    float(cashflow.cash_paid_goods) if cashflow.cash_paid_goods else 0,
                    float(cashflow.cash_paid_employees) if cashflow.cash_paid_employees else 0,
                    float(cashflow.cash_paid_taxes) if cashflow.cash_paid_taxes else 0,
                    0,  # 经营活动现金流出小计
                    float(cashflow.net_cash_operating) if cashflow.net_cash_operating else 0,
                    float(cashflow.cash_received_investments)
                    if cashflow.cash_received_investments
                    else 0,
                    0,
                    0,  # 取得投资收益, 投资活动现金流入小计
                    0,
                    0,  # 购建固定资产, 投资支付
                    0,  # 投资活动现金流出小计
                    float(cashflow.net_cash_investing) if cashflow.net_cash_investing else 0,
                    0,
                    0,
                    0,
                    0,
                    0,
                    0,  # 筹资活动
                    float(cashflow.net_cash_financing) if cashflow.net_cash_financing else 0,
                    float(cashflow.net_cash_increase) if cashflow.net_cash_increase else 0,
                    float(cashflow.cash_end_period) if cashflow.cash_end_period else 0,
                ]
                for col, value in enumerate(row_data, 1):
                    ws_cashflow.cell(row=row_cashflow, column=col, value=value).border = thin_border
                row_cashflow += 1

    # Adjust column widths
    for ws in [ws_balance, ws_income, ws_cashflow]:
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width

    # Save to bytes
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    # Return as streaming response
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=financial_data.xlsx"},
    )
